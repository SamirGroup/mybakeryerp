import io
import json
import requests as http_requests
from datetime import date, timedelta, datetime, time
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from urllib.parse import urlencode

from .models import AdvancePayment, DailyReport, Employee, Position, Shift, Attendance, EmployeePhoto, FaceIDLog


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _can_access(user, *roles):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=roles).exists()


def _can_access_branch(user, branch_id=None):
    if user.is_superuser:
        return True
    if 'branch_admin' in user.groups.values_list('name', flat=True):
        try:
            user_branch = user.profile.branch
            if branch_id and user_branch.id != branch_id:
                return False
            return True
        except Exception:
            return False
    return False


def _parse_decimal(val, default=None):
    val = (val or '').strip()
    if not val:
        return default
    try:
        return Decimal(val.replace(',', '.'))
    except InvalidOperation:
        return default


def _date_range(request):
    period = request.GET.get('period', 'today')
    today = timezone.localdate()
    if period == 'week':
        return today - timedelta(days=6), today
    if period == 'custom':
        try:
            df = date.fromisoformat(request.GET.get('date_from', ''))
            dt = date.fromisoformat(request.GET.get('date_to', ''))
            return df, dt
        except ValueError:
            pass
    return today, today


def _send_telegram(message: str, employee=None):
    """Telegram bot orqali xabar yuborish.
    Tartib: 1) Xodimning filial Telegram → 2) Global DB → 3) settings.py
    """
    configs = []

    # 1. Xodim filialni aniqlash
    if employee and employee.branch_id:
        try:
            from core.models import BranchTelegramSettings
            br_tg = BranchTelegramSettings.get_for_branch(employee.branch)
            if br_tg.is_active and br_tg.bot_token and br_tg.chat_id:
                configs.append((br_tg.bot_token, br_tg.chat_id))
        except Exception:
            pass

    # 2. Global DB sozlamalari
    if not configs:
        try:
            from core.models import TelegramSettings
            cfg = TelegramSettings.get()
            if cfg.is_active and cfg.bot_token and cfg.chat_id:
                configs.append((cfg.bot_token, cfg.chat_id))
        except Exception:
            pass

    # 3. settings.py fallback
    if not configs:
        token = getattr(settings, 'TELEGRAM_BOT_TOKEN', '')
        chat_id = getattr(settings, 'TELEGRAM_CHAT_ID', '')
        if token and chat_id:
            configs.append((token, chat_id))

    for token, chat_id in configs:
        try:
            http_requests.post(
                f'https://api.telegram.org/bot{token}/sendMessage',
                json={'chat_id': chat_id, 'text': message, 'parse_mode': 'HTML'},
                timeout=5,
            )
        except Exception:
            pass


# ─── HR Dashboard ─────────────────────────────────────────────────────────────

@login_required
def hr_dashboard(request):
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return redirect('dashboard')

    if _can_access_branch(request.user) and not request.user.is_superuser:
        try:
            user_branch = request.user.profile.branch
            employees = Employee.objects.select_related('shift', 'branch').filter(branch=user_branch)
        except Exception:
            employees = Employee.objects.none()
    else:
        employees = Employee.objects.select_related('shift').all()

    shifts = Shift.objects.all().order_by('name')
    positions = Position.objects.all().order_by('name')
    advances = AdvancePayment.objects.select_related('employee').order_by('-date')[:15]

    date_from, date_to = _date_range(request)

    sort_by = request.GET.get('sort', 'name')
    sort_map = {
        'name': 'name',
        'position': 'position',
        'shift': 'shift__name',
        'date_joined': 'date_joined',
        'target': '-daily_target',
    }
    employees = employees.order_by(sort_map.get(sort_by, 'name'))

    shift_filter = request.GET.get('shift_filter', '')
    if shift_filter:
        employees = employees.filter(shift_id=shift_filter)

    pos_filter = request.GET.get('pos_filter', '')
    if pos_filter:
        employees = employees.filter(position__icontains=pos_filter)

    if request.GET.get('export') == 'excel':
        return _export_hr_excel(employees, date_from, date_to)

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add_shift':
            name = request.POST.get('shift_name', '').strip()
            start = request.POST.get('shift_start', '')
            end = request.POST.get('shift_end', '')
            if name and start and end:
                Shift.objects.get_or_create(name=name, defaults={'start_time': start, 'end_time': end})
                messages.success(request, f"Smena '{name}' qo'shildi.")
            else:
                messages.error(request, "Smena nomi, boshlanish va tugash vaqti majburiy.")

        elif action == 'delete_shift':
            sid = request.POST.get('shift_id')
            try:
                Shift.objects.get(id=sid).delete()
                messages.success(request, "Smena o'chirildi.")
            except Shift.DoesNotExist:
                messages.error(request, "Smena topilmadi.")

        elif action == 'add_position':
            name = request.POST.get('position_name', '').strip()
            desc = request.POST.get('position_desc', '').strip()
            if name:
                _, created = Position.objects.get_or_create(name=name, defaults={'description': desc})
                if created:
                    messages.success(request, f"Lavozim '{name}' qo'shildi.")
                else:
                    messages.warning(request, f"'{name}' lavozimi allaqachon mavjud.")
            else:
                messages.error(request, "Lavozim nomi majburiy.")

        elif action == 'edit_position':
            pid = request.POST.get('position_id')
            try:
                pos = Position.objects.get(id=pid)
                new_name = request.POST.get('position_name', '').strip()
                if new_name:
                    pos.name = new_name
                pos.description = request.POST.get('position_desc', '').strip()
                pos.save()
                messages.success(request, f"Lavozim '{pos.name}' yangilandi.")
            except Position.DoesNotExist:
                messages.error(request, "Lavozim topilmadi.")

        elif action == 'delete_position':
            pid = request.POST.get('position_id')
            try:
                Position.objects.get(id=pid).delete()
                messages.success(request, "Lavozim o'chirildi.")
            except Position.DoesNotExist:
                messages.error(request, "Lavozim topilmadi.")

        elif action == 'add_employee':
            name = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()
            position = request.POST.get('position', '').strip()
            date_joined = request.POST.get('date_joined', '')
            shift_id = request.POST.get('shift_id', '')
            is_piecework = request.POST.get('is_piecework') == 'on'
            base_salary = request.POST.get('base_salary', 0) or 0
            piecework_rate = request.POST.get('piecework_rate', 0) or 0
            daily_target = request.POST.get('daily_target', 0) or 0
            face_photos = request.FILES.getlist('face_photos')
            branch_id = request.POST.get('branch_id', '')

            if name and position and date_joined:
                shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
                emp, created = Employee.objects.get_or_create(
                    name=name, phone=phone,
                    defaults={
                        'position': position,
                        'date_joined': date_joined,
                        'is_piecework': is_piecework,
                        'base_salary': base_salary,
                        'piecework_rate': piecework_rate,
                        'daily_target': daily_target,
                        'status': 'active',
                        'shift': shift,
                    }
                )
                if created:
                    uid = request.POST.get('user_account_id', '').strip()
                    if uid:
                        try:
                            oid = int(uid)
                            if not Employee.objects.filter(user_account_id=oid).exists():
                                emp.user_account_id = oid
                                emp.save(update_fields=['user_account_id'])
                        except ValueError:
                            pass

                    if branch_id:
                        try:
                            from branches.models import Branch
                            emp.branch = Branch.objects.get(id=branch_id)
                            emp.save(update_fields=['branch'])
                        except Exception:
                            pass
                    elif _can_access_branch(request.user) and not request.user.is_superuser:
                        try:
                            emp.branch = request.user.profile.branch
                            emp.save(update_fields=['branch'])
                        except Exception:
                            pass

                    if face_photos:
                        emp.face_id_enrolled = True
                        emp.save(update_fields=['face_id_enrolled'])
                        for i, photo_file in enumerate(face_photos[:5]):
                            EmployeePhoto.objects.create(
                                employee=emp,
                                photo=photo_file,
                                is_primary=(i == 0)
                            )

                    messages.success(request, f"Xodim '{name}' qo'shildi.")
                else:
                    messages.warning(request, "Bu xodim allaqachon mavjud.")
            else:
                messages.error(request, "Ism, lavozim va sana majburiy.")

        elif action == 'edit_employee':
            eid = request.POST.get('employee_id')
            try:
                emp = Employee.objects.get(id=eid)
                emp.name = request.POST.get('name', emp.name).strip()
                emp.phone = request.POST.get('phone', emp.phone).strip()
                emp.position = request.POST.get('position', emp.position).strip()
                emp.date_joined = request.POST.get('date_joined', emp.date_joined)
                shift_id = request.POST.get('shift_id', '')
                emp.shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
                emp.is_piecework = request.POST.get('is_piecework') == 'on'
                emp.base_salary = request.POST.get('base_salary', emp.base_salary) or emp.base_salary
                emp.piecework_rate = request.POST.get('piecework_rate', emp.piecework_rate) or emp.piecework_rate
                emp.daily_target = request.POST.get('daily_target', emp.daily_target) or emp.daily_target
                new_status = request.POST.get('status', '').strip()
                if new_status in dict(Employee.STATUS_CHOICES):
                    emp.status = new_status
                uid = request.POST.get('user_account_id', '').strip()
                if uid:
                    try:
                        oid = int(uid)
                        if Employee.objects.filter(user_account_id=oid).exclude(pk=emp.pk).exists():
                            messages.error(request, "Bu login boshqa xodimga biriktirilgan.")
                        else:
                            emp.user_account_id = oid
                    except ValueError:
                        pass
                else:
                    emp.user_account_id = None
                branch_id = request.POST.get('branch_id', '')
                if branch_id:
                    try:
                        from branches.models import Branch
                        emp.branch = Branch.objects.get(id=branch_id)
                    except Exception:
                        pass
                emp.save()
                messages.success(request, f"{emp.name} yangilandi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        elif action == 'change_status':
            eid = request.POST.get('employee_id')
            new_status = request.POST.get('new_status')
            try:
                emp = Employee.objects.get(id=eid)
                emp.status = new_status
                emp.save()
                messages.success(request, f"{emp.name} holati o'zgartirildi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        elif action == 'delete_employee':
            eid = request.POST.get('employee_id')
            try:
                Employee.objects.get(id=eid).delete()
                messages.success(request, "Xodim o'chirildi.")
            except Employee.DoesNotExist:
                messages.error(request, "Topilmadi.")

        elif action == 'add_daily_report':
            eid = request.POST.get('employee_id')
            report_date = request.POST.get('report_date', str(timezone.localdate()))
            shift_id = request.POST.get('shift_id', '')
            check_in = request.POST.get('check_in') or None
            check_out = request.POST.get('check_out') or None
            units = int(request.POST.get('units_produced', 0) or 0)
            notes = request.POST.get('notes', '')
            hours_expected = _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8')
            hours_present = _parse_decimal(request.POST.get('hours_present'))
            hours_absent = _parse_decimal(request.POST.get('hours_absent'))
            hours_left_early = _parse_decimal(request.POST.get('hours_left_early'))
            was_present = request.POST.get('attendance', 'present') == 'present'
            try:
                emp = Employee.objects.get(id=eid)
                shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
                DailyReport.objects.update_or_create(
                    employee=emp, date=report_date, shift=shift,
                    defaults={
                        'check_in': check_in,
                        'check_out': check_out,
                        'units_produced': units,
                        'notes': notes,
                        'hours_expected': hours_expected,
                        'hours_present': hours_present,
                        'hours_absent': hours_absent,
                        'hours_left_early': hours_left_early,
                        'was_present': was_present,
                    },
                )
                messages.success(request, f"{emp.name} uchun hisobot saqlandi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        elif 'give_advance' in request.POST:
            emp_id = request.POST.get('employee')
            amount = request.POST.get('amount')
            try:
                emp = Employee.objects.get(id=emp_id)
                AdvancePayment.objects.create(employee=emp, amount=amount)
                messages.success(request, f"{emp.name} ga {amount} UZS avans berildi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        params = {
            'period': request.GET.get('period', 'today'),
            'sort': sort_by,
            'shift_filter': shift_filter,
            'pos_filter': pos_filter,
            'date_from': request.GET.get('date_from', ''),
            'date_to': request.GET.get('date_to', ''),
        }
        return redirect('/hr/?' + urlencode(params))

    # Daily reports
    if _can_access_branch(request.user) and not request.user.is_superuser:
        try:
            user_branch = request.user.profile.branch
            daily_reports = (
                DailyReport.objects
                .select_related('employee', 'shift')
                .filter(date__gte=date_from, date__lte=date_to, employee__branch=user_branch)
                .order_by('-date', 'employee__name')
            )
            face_logs = FaceIDLog.objects.select_related('employee').filter(
                timestamp__date=timezone.localdate(),
                employee__branch=user_branch
            ).order_by('-timestamp')[:20]
            attendances = Attendance.objects.select_related('employee').filter(
                date=timezone.localdate(),
                employee__branch=user_branch
            )
            enrolled_employees = Employee.objects.filter(face_id_enrolled=True, branch=user_branch)
        except Exception:
            daily_reports = DailyReport.objects.none()
            face_logs = FaceIDLog.objects.none()
            attendances = Attendance.objects.none()
            enrolled_employees = Employee.objects.none()
    else:
        daily_reports = (
            DailyReport.objects
            .select_related('employee', 'shift')
            .filter(date__gte=date_from, date__lte=date_to)
            .order_by('-date', 'employee__name')
        )
        face_logs = FaceIDLog.objects.select_related('employee').filter(
            timestamp__date=timezone.localdate()
        ).order_by('-timestamp')[:20]
        attendances = Attendance.objects.select_related('employee').filter(date=timezone.localdate())
        enrolled_employees = Employee.objects.filter(face_id_enrolled=True)

    late_arrivals = attendances.filter(late_minutes__gt=0)
    seller_users = User.objects.filter(is_active=True).order_by('username')

    from branches.models import Branch
    branches = Branch.objects.filter(is_active=True).order_by('name')

    employee_edit_payload = [
        {
            'id': e.id,
            'name': e.name,
            'phone': e.phone or '',
            'position': e.position,
            'date_joined': str(e.date_joined),
            'shift_id': e.shift_id,
            'status': e.status,
            'is_piecework': bool(e.is_piecework),
            'base_salary': str(e.base_salary),
            'piecework_rate': str(e.piecework_rate),
            'daily_target': e.daily_target,
            'user_account_id': e.user_account_id,
            'branch_id': e.branch_id,
        }
        for e in employees
    ]

    context = {
        'employees': employees,
        'active_employees': employees.filter(status='active'),
        'seller_users': seller_users,
        'employee_edit_payload': json.dumps(employee_edit_payload),
        'shifts': shifts,
        'positions': positions,
        'advances': advances,
        'daily_reports': daily_reports,
        'today': timezone.localdate(),
        'date_from': date_from,
        'date_to': date_to,
        'period': request.GET.get('period', 'today'),
        'sort_by': sort_by,
        'shift_filter': shift_filter,
        'pos_filter': pos_filter,
        'face_logs': face_logs,
        'late_arrivals': late_arrivals,
        'enrolled_employees': enrolled_employees,
        'attendances': attendances,
        'branches': branches,
        'telegram_configured': bool(getattr(settings, 'TELEGRAM_BOT_TOKEN', '')),
    }
    return render(request, 'hr.html', context)


# ─── Employee Report ───────────────────────────────────────────────────────────

@login_required
def employee_report(request, emp_id):
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return redirect('dashboard')
    try:
        emp = Employee.objects.select_related('shift').get(id=emp_id)
    except Employee.DoesNotExist:
        return redirect('hr_dashboard')

    today = timezone.localdate()
    period = request.GET.get('period', 'month')

    if period == 'today':
        date_from = date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'custom':
        try:
            date_from = date.fromisoformat(request.GET.get('date_from', ''))
            date_to   = date.fromisoformat(request.GET.get('date_to', ''))
        except ValueError:
            date_from = today.replace(day=1)
            date_to   = today
    else:
        date_from = today.replace(day=1)
        date_to   = today

    if request.method == 'POST':
        action   = request.POST.get('action', '')
        shift_id = request.POST.get('shift_id', '')
        shift    = Shift.objects.filter(id=shift_id).first() if shift_id else None
        was_present    = request.POST.get('was_present') == '1'
        absence_reason = '' if was_present else request.POST.get('absence_reason', '')
        defaults = {
            'check_in':       request.POST.get('check_in') or None,
            'check_out':      request.POST.get('check_out') or None,
            'was_present':    was_present,
            'absence_reason': absence_reason,
            'units_produced': int(request.POST.get('units_produced', 0) or 0),
            'hours_expected': _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8'),
            'hours_present':  _parse_decimal(request.POST.get('hours_present')),
            'notes':          request.POST.get('notes', ''),
        }
        if action == 'edit_report':
            rid = request.POST.get('report_id')
            try:
                r = DailyReport.objects.get(id=rid, employee=emp)
                for k, v in defaults.items():
                    setattr(r, k, v)
                r.save()
                messages.success(request, f"{r.date} hisoboti yangilandi.")
            except DailyReport.DoesNotExist:
                messages.error(request, "Hisobot topilmadi.")
        elif action == 'add_report':
            rdate = request.POST.get('report_date', str(today))
            DailyReport.objects.update_or_create(
                employee=emp, date=rdate, shift=shift,
                defaults=defaults
            )
            messages.success(request, "Hisobot saqlandi.")
        return redirect(f'/hr/employee/{emp_id}/report/?period={period}&date_from={date_from}&date_to={date_to}')

    reports = (
        DailyReport.objects
        .filter(employee=emp, date__gte=date_from, date__lte=date_to)
        .order_by('date')
    )

    present_days     = reports.filter(was_present=True).count()
    absent_days      = reports.filter(was_present=False).count()
    absent_sick      = reports.filter(was_present=False, absence_reason='sick').count()
    absent_personal  = reports.filter(was_present=False, absence_reason='personal').count()
    absent_approved  = reports.filter(was_present=False, absence_reason='approved').count()
    absent_no_reason = reports.filter(was_present=False, absence_reason='no_reason').count()

    total_units = sum((r.units_produced or 0) + (r.units_from_sales or 0) for r in reports)
    total_hours = round(sum(float(r.hours_present or 0) for r in reports), 2)
    total_days  = (date_to - date_from).days + 1

    if emp.is_piecework:
        salary_calc = Decimal(str(total_units)) * (emp.piecework_rate or Decimal('0'))
        salary_type = 'piecework'
    else:
        daily_rate  = (emp.base_salary or Decimal('0')) / Decimal(str(total_days)) if total_days else Decimal('0')
        salary_calc = daily_rate * Decimal(str(present_days))
        salary_type = 'daily'

    advances      = AdvancePayment.objects.filter(employee=emp, date__gte=date_from, date__lte=date_to)
    total_advance = sum(a.amount for a in advances) or Decimal('0')
    net_salary    = salary_calc - total_advance

    daily_target_status = calculate_daily_target_status(emp, today) if emp.is_piecework else None

    # ── KPI Statistikasi ────────────────────────────────────────────────
    attendance_pct = round((present_days / total_days * 100), 1) if total_days > 0 else 0
    target_fulfillment_pct = 0
    if emp.is_piecework and emp.daily_target and present_days > 0:
        avg_daily_units = total_units / present_days
        target_fulfillment_pct = round(min(avg_daily_units / emp.daily_target * 100, 100), 1)

    # Davomat qoidabuzarliklari
    from .models import Attendance
    late_records = Attendance.objects.filter(
        employee=emp, date__gte=date_from, date__lte=date_to, late_minutes__gt=0
    )
    late_count = late_records.count()
    avg_late_minutes = 0
    if late_count:
        avg_late_minutes = round(late_records.aggregate(a=Sum('late_minutes'))['a'] / late_count)

    # O'rtacha kunlik ishlab chiqarish
    avg_daily_units = round(total_units / present_days, 1) if present_days > 0 else 0
    avg_daily_hours = round(total_hours / present_days, 1) if present_days > 0 else 0

    # Savdo qo'shish (agar piecework va sales ga ulangan bo'lsa)
    units_from_sales = sum(r.units_from_sales or 0 for r in reports)
    units_from_production = sum(r.units_produced or 0 for r in reports)

    # KPI ball (0-100): davomat 40% + nagruzka 40% + kechikmaslik 20%
    kpi_attendance  = min(attendance_pct, 100) * 0.40
    kpi_target      = min(target_fulfillment_pct, 100) * 0.40
    kpi_punctuality = max(0, 100 - (late_count * 5)) * 0.20  # har kechikish uchun -5 ball
    kpi_score = round(kpi_attendance + kpi_target + kpi_punctuality, 1)

    kpi = {
        'score': kpi_score,
        'attendance_pct': attendance_pct,
        'target_pct': target_fulfillment_pct,
        'late_count': late_count,
        'avg_late_min': avg_late_minutes,
        'avg_daily_units': avg_daily_units,
        'avg_daily_hours': avg_daily_hours,
        'units_from_sales': units_from_sales,
        'units_from_production': units_from_production,
        'grade': 'A' if kpi_score >= 90 else 'B' if kpi_score >= 75 else 'C' if kpi_score >= 60 else 'D',
        'grade_color': '#27ae60' if kpi_score >= 90 else '#f39c12' if kpi_score >= 75 else '#e67e22' if kpi_score >= 60 else '#e74c3c',
    }

    context = {
        'emp': emp,
        'reports': reports,
        'shifts': Shift.objects.all().order_by('name'),
        'date_from': date_from,
        'date_to': date_to,
        'period': period,
        'today': today,
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'absent_sick': absent_sick,
        'absent_personal': absent_personal,
        'absent_approved': absent_approved,
        'absent_no_reason': absent_no_reason,
        'total_units': total_units,
        'total_hours': total_hours,
        'salary_calc': salary_calc,
        'salary_type': salary_type,
        'total_advance': total_advance,
        'net_salary': net_salary,
        'advances': advances,
        'absence_reasons': DailyReport.ABSENCE_REASON_CHOICES,
        'daily_target_status': daily_target_status,
        'kpi': kpi,
    }
    return render(request, 'hr_employee_report.html', context)


# ─── Positions Export/Import ───────────────────────────────────────────────────

@login_required
def positions_export_json(request):
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return redirect('dashboard')
    data = [
        {'name': p.name, 'description': p.description}
        for p in Position.objects.order_by('name')
    ]
    resp = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json'
    )
    resp['Content-Disposition'] = 'attachment; filename="positions.json"'
    return resp


@login_required
def positions_import_json(request):
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('hr_dashboard')
    f = request.FILES.get('positions_json_file')
    if not f:
        messages.error(request, "JSON fayl tanlanmadi.")
        return redirect('/hr/?tab=positions')
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception as e:
        messages.error(request, f"JSON o'qishda xato: {e}")
        return redirect('/hr/?tab=positions')
    created = updated = 0
    for entry in data:
        name = str(entry.get('name', '')).strip()
        if not name:
            continue
        desc = str(entry.get('description', '')).strip()
        pos, is_new = Position.objects.get_or_create(name=name, defaults={'description': desc})
        if not is_new:
            pos.description = desc
            pos.save(update_fields=['description'])
            updated += 1
        else:
            created += 1
    messages.success(request, f"Import tugadi: {created} yangi, {updated} yangilandi.")
    return redirect('/hr/?tab=positions')


def _export_hr_excel(employees, date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xodimlar"
    fill = PatternFill("solid", fgColor="D4A373")  # #D4A373 asosiy rang
    bold = Font(bold=True)
    headers = ['Ism', 'Lavozim', 'Smena', 'Telefon', 'Qo\'shilgan sana', 'Holat', 'Nagruzka (kun)', 'Ishbay stavka']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    for row_idx, emp in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp.name)
        ws.cell(row=row_idx, column=2, value=emp.position)
        ws.cell(row=row_idx, column=3, value=emp.shift.name if emp.shift else '')
        ws.cell(row=row_idx, column=4, value=emp.phone)
        ws.cell(row=row_idx, column=5, value=str(emp.date_joined))
        ws.cell(row=row_idx, column=6, value=emp.get_status_display())
        ws.cell(row=row_idx, column=7, value=emp.daily_target)
        ws.cell(row=row_idx, column=8, value=float(emp.piecework_rate))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="employees_{date_from}_{date_to}.xlsx"'
    return resp


# ─── Kunlik Nagruzka ───────────────────────────────────────────────────────────

def calculate_daily_target_status(employee, report_date=None):
    """Xodimning kunlik nagruzka bajarilishini hisoblash."""
    if not report_date:
        report_date = timezone.localdate()
    try:
        report = DailyReport.objects.get(employee=employee, date=report_date)
    except DailyReport.DoesNotExist:
        return None

    target = employee.daily_target or 0
    total_units = report.piecework_units_total

    if target <= 0:
        return {
            'target': 0,
            'completed': total_units,
            'percentage': 100 if total_units > 0 else 0,
            'status': 'no_target',
            'status_display': 'Nagruzka belgilanmagan',
            'earnings': report.estimated_piecework_earn,
            'bonus': Decimal('0'),
            'total_payment': report.estimated_piecework_earn,
            'piecework_rate': employee.piecework_rate,
        }

    percentage = min((total_units / target) * 100, 100)

    if total_units >= target:
        status = 'completed'
        status_display = 'Bajarildi ✅'
    elif total_units >= target * 0.75:
        status = 'partial'
        status_display = 'Yarim bajarildi ⚠️'
    else:
        status = 'failed'
        status_display = 'Bajarilmadi ❌'

    earnings = report.estimated_piecework_earn
    bonus = earnings * Decimal('0.10') if status == 'completed' else Decimal('0')

    return {
        'target': target,
        'completed': total_units,
        'percentage': round(percentage, 1),
        'status': status,
        'status_display': status_display,
        'earnings': earnings,
        'bonus': bonus,
        'total_payment': earnings + bonus,
        'piecework_rate': employee.piecework_rate,
    }


@login_required
def daily_target_api(request, emp_id):
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return JsonResponse({'error': "Ruxsat yo'q"}, status=403)
    try:
        emp = Employee.objects.get(id=emp_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Xodim topilmadi'}, status=404)

    date_str = request.GET.get('date')
    report_date = timezone.localdate()
    if date_str:
        try:
            report_date = date.fromisoformat(date_str)
        except ValueError:
            pass

    result = calculate_daily_target_status(emp, report_date)
    if result is None:
        return JsonResponse({'error': 'Hisobot topilmadi', 'employee': emp.name, 'date': str(report_date)}, status=404)

    return JsonResponse({
        'success': True,
        'employee': emp.name,
        'position': emp.position,
        'date': str(report_date),
        'target': result['target'],
        'completed': result['completed'],
        'percentage': result['percentage'],
        'status': result['status'],
        'status_display': result['status_display'],
        'earnings': float(result['earnings']),
        'bonus': float(result['bonus']),
        'total_payment': float(result['total_payment']),
        'piecework_rate': float(result['piecework_rate']),
    })


# ─── Face ID Views ─────────────────────────────────────────────────────────────

@login_required
def face_id_check_in(request):
    """Face ID orqali kirish - API endpoint"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    employee_id = request.POST.get('employee_id')
    confidence = float(request.POST.get('confidence', 0))
    snapshot = request.FILES.get('snapshot')

    try:
        emp = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Xodim topilmadi'}, status=404)

    if not emp.face_id_enrolled:
        return JsonResponse({'error': 'Xodim Face ID tizimiga ulanmagan'}, status=400)

    today = timezone.localdate()
    now = timezone.now()
    current_time = now.time()

    shift_start = emp.shift.start_time if emp.shift else None
    expected_check_in = shift_start

    late_minutes = 0
    is_late = False
    if shift_start and current_time > shift_start:
        diff = datetime.combine(today, current_time) - datetime.combine(today, shift_start)
        late_minutes = int(diff.total_seconds() / 60)
        is_late = late_minutes > 0

    log = FaceIDLog.objects.create(
        employee=emp,
        action='check_in',
        confidence=confidence,
        is_late=is_late,
        late_minutes=late_minutes,
        shift_start_time=shift_start,
    )
    if snapshot:
        log.snapshot = snapshot
        log.save(update_fields=['snapshot'])

    Attendance.objects.update_or_create(
        employee=emp,
        date=today,
        defaults={
            'check_in': current_time,
            'check_in_method': 'face_id',
            'late_minutes': late_minutes,
            'expected_check_in': expected_check_in,
        }
    )

    time_str = current_time.strftime('%H:%M')
    shift_name = emp.shift.name if emp.shift else 'Smenasiz'
    if is_late:
        tg_msg = (
            f"⚠️ <b>Kechikish!</b>\n"
            f"👤 {emp.name} — {emp.position}\n"
            f"🕐 Keldi: {time_str} ({late_minutes} daqiqa kech)\n"
            f"📋 Smena: {shift_name}"
        )
        ui_msg = f"⚠️ {emp.name} {late_minutes} daqiqa kechikdi"
    else:
        tg_msg = (
            f"✅ <b>Keldi</b>\n"
            f"👤 {emp.name} — {emp.position}\n"
            f"🕐 Vaqt: {time_str}\n"
            f"📋 Smena: {shift_name}"
        )
        ui_msg = f"✅ {emp.name} keldi"

    _send_telegram(tg_msg, employee=emp)

    return JsonResponse({
        'success': True,
        'message': ui_msg,
        'employee_name': emp.name,
        'position': emp.position,
        'check_in_time': time_str,
        'is_late': is_late,
        'late_minutes': late_minutes,
        'shift': shift_name,
    })


@login_required
def face_id_check_out(request):
    """Face ID orqali chiqish - API endpoint"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    employee_id = request.POST.get('employee_id')
    confidence = float(request.POST.get('confidence', 0))
    snapshot = request.FILES.get('snapshot')

    try:
        emp = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Xodim topilmadi'}, status=404)

    today = timezone.localdate()
    now = timezone.now()
    current_time = now.time()

    log = FaceIDLog.objects.create(
        employee=emp,
        action='check_out',
        confidence=confidence,
        shift_start_time=emp.shift.start_time if emp.shift else None,
    )
    if snapshot:
        log.snapshot = snapshot
        log.save(update_fields=['snapshot'])

    try:
        att = Attendance.objects.get(employee=emp, date=today)
        att.check_out = current_time
        att.check_out_method = 'face_id'
        att.save()
    except Attendance.DoesNotExist:
        Attendance.objects.create(
            employee=emp,
            date=today,
            check_out=current_time,
            check_out_method='face_id',
        )

    time_str = current_time.strftime('%H:%M')
    tg_msg = (
        f"👋 <b>Ketdi</b>\n"
        f"👤 {emp.name} — {emp.position}\n"
        f"🕐 Vaqt: {time_str}"
    )
    _send_telegram(tg_msg, employee=emp)

    return JsonResponse({
        'success': True,
        'message': f"👋 {emp.name} ketdi",
        'employee_name': emp.name,
        'position': emp.position,
        'check_out_time': time_str,
    })


@login_required
def face_id_enroll(request):
    """Xodimni Face ID tizimiga ulash"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    employee_id = request.POST.get('employee_id')
    try:
        emp = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Xodim topilmadi'}, status=404)

    photos = request.FILES.getlist('photos')
    if not photos:
        return JsonResponse({'error': "Kamida bitta rasm yuklanishi kerak"}, status=400)

    emp.face_id_enrolled = True
    emp.save(update_fields=['face_id_enrolled'])

    saved = 0
    existing_count = EmployeePhoto.objects.filter(employee=emp).count()
    for i, photo_file in enumerate(photos):
        EmployeePhoto.objects.create(
            employee=emp,
            photo=photo_file,
            is_primary=(existing_count == 0 and i == 0)
        )
        saved += 1

    return JsonResponse({
        'success': True,
        'message': f"{emp.name} Face ID tizimiga ulandi",
        'employee_id': emp.id,
        'photos_count': saved,
    })


@login_required
def face_dashboard(request):
    """Face ID monitoring dashboard"""
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return redirect('dashboard')

    today = timezone.localdate()

    if _can_access_branch(request.user) and not request.user.is_superuser:
        try:
            user_branch = request.user.profile.branch
            face_logs = FaceIDLog.objects.select_related('employee').filter(
                timestamp__date=today,
                employee__branch=user_branch
            ).order_by('-timestamp')[:50]
            attendances = Attendance.objects.select_related('employee', 'employee__shift').filter(
                date=today,
                employee__branch=user_branch
            )
            enrolled_employees = Employee.objects.filter(face_id_enrolled=True, branch=user_branch)
        except Exception:
            face_logs = FaceIDLog.objects.none()
            attendances = Attendance.objects.none()
            enrolled_employees = Employee.objects.none()
    else:
        face_logs = FaceIDLog.objects.select_related('employee').filter(
            timestamp__date=today
        ).order_by('-timestamp')[:50]
        attendances = Attendance.objects.select_related('employee', 'employee__shift').filter(date=today)
        enrolled_employees = Employee.objects.filter(face_id_enrolled=True)

    late_arrivals = attendances.filter(late_minutes__gt=0)

    context = {
        'face_logs': face_logs,
        'attendances': attendances,
        'late_arrivals': late_arrivals,
        'enrolled_employees': enrolled_employees,
        'today': today,
    }
    return render(request, 'face_dashboard.html', context)


@login_required
def face_id_camera(request):
    """Face ID kamera sahifasi"""
    if not _can_access(request.user, 'hr', 'branch_admin'):
        return redirect('dashboard')

    today = timezone.localdate()

    if _can_access_branch(request.user) and not request.user.is_superuser:
        try:
            user_branch = request.user.profile.branch
            enrolled_employees = Employee.objects.filter(face_id_enrolled=True, branch=user_branch).select_related('shift')
            all_employees = Employee.objects.filter(status='active', branch=user_branch).select_related('shift')
        except Exception:
            enrolled_employees = Employee.objects.none()
            all_employees = Employee.objects.none()
    else:
        enrolled_employees = Employee.objects.filter(face_id_enrolled=True).select_related('shift')
        all_employees = Employee.objects.filter(status='active').select_related('shift')

    context = {
        'enrolled_employees': enrolled_employees,
        'all_employees': all_employees,
        'today': today,
        'telegram_configured': bool(getattr(settings, 'TELEGRAM_BOT_TOKEN', '')),
    }
    return render(request, 'hr_face_camera.html', context)


@login_required
def api_check_in(request):
    """API: Qo'lda kirish/chiqish"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': "JSON xato"}, status=400)

    employee_id = data.get('employee_id')
    action = data.get('action')

    try:
        emp = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Xodim topilmadi'}, status=404)

    today = timezone.localdate()
    now = timezone.now()
    current_time = now.time()
    time_str = current_time.strftime('%H:%M')

    if action == 'check_in':
        late_minutes = 0
        is_late = False
        shift_start = emp.shift.start_time if emp.shift else None

        if shift_start and current_time > shift_start:
            diff = datetime.combine(today, current_time) - datetime.combine(today, shift_start)
            late_minutes = int(diff.total_seconds() / 60)
            is_late = late_minutes > 0

        Attendance.objects.update_or_create(
            employee=emp,
            date=today,
            defaults={
                'check_in': current_time,
                'check_in_method': 'manual',
                'late_minutes': late_minutes,
                'expected_check_in': shift_start,
            }
        )

        shift_name = emp.shift.name if emp.shift else 'Smenasiz'
        if is_late:
            tg_msg = f"⚠️ <b>{emp.name}</b> {late_minutes} daqiqa kechikib keldi\n🕐 {time_str} — {emp.position}\n📋 {shift_name}"
            ui_msg = f"⚠️ {emp.name} {late_minutes} daqiqa kechikdi"
        else:
            tg_msg = f"✅ <b>{emp.name}</b> keldi\n🕐 {time_str} — {emp.position}\n📋 {shift_name}"
            ui_msg = f"✅ {emp.name} keldi"

        _send_telegram(tg_msg)
        return JsonResponse({'success': True, 'message': ui_msg, 'employee_name': emp.name,
                             'check_in_time': time_str, 'is_late': is_late, 'late_minutes': late_minutes})

    elif action == 'check_out':
        try:
            att = Attendance.objects.get(employee=emp, date=today)
            att.check_out = current_time
            att.check_out_method = 'manual'
            att.save()
        except Attendance.DoesNotExist:
            Attendance.objects.create(employee=emp, date=today, check_out=current_time, check_out_method='manual')

        tg_msg = f"👋 <b>{emp.name}</b> ketdi\n🕐 {time_str} — {emp.position}"
        _send_telegram(tg_msg)
        return JsonResponse({'success': True, 'message': f"👋 {emp.name} ketdi",
                             'employee_name': emp.name, 'check_out_time': time_str})

    return JsonResponse({"error": "Noto'g'ri harakat"}, status=400)
